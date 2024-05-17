import pandas as pd
import openpyxl
import os

wb=openpyxl.Workbook()
ws=wb.active
#path
excelname = 'C:/Users/15920/Desktop/jicheng/data/val/valDataExcel/R5S.xlsx'
wb.save(excelname)
# read txt
col = 1
def get_file_name_1(file_path):
    file_names = os.listdir(file_path)
    return file_names
file_path = 'C:/Users/15920/Desktop/jicheng/data/val/R5S'
file_names = get_file_name_1(file_path)
# print(file_names)
print(len(file_names))
for j in file_names:
    fopen = open(file_path+'/'+j, 'r',encoding='utf-8')
    lines = fopen.readlines()
    #写excel
    file = openpyxl.load_workbook(excelname)
    sheet = file.active
    row=1
    cnt = 0
    for line in lines:
        if cnt < 5:
            cnt+=1
            continue
        #换行
        line = line.strip('\n')
        line = line.split(',')
        for index in range(1,len(line),2):
            sheet.cell(row, col, line[index])
            row += 1
        # 行数递增
    col += 1
    file.save(excelname)
file.close()

xl=openpyxl.load_workbook('C:/Users/15920/Desktop/jicheng/data/val/valDataExcel/R5S.xlsx')
xl_sheet_names=xl.get_sheet_names()
xl_sheet=xl.get_sheet_by_name(xl_sheet_names[0])
print(xl_sheet.max_row)
print(xl_sheet.max_column)

